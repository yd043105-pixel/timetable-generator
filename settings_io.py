# -*- coding: utf-8 -*-
"""입력 설정(비수업·교사불가·유사그룹·기본설정)을 엑셀로 저장/불러오기.

시수표와 별개로, 손으로 맞춘 조건들을 .xlsx 로 보관했다가 다시 적용한다.
사람이 직접 열어 확인·수정할 수 있도록 시트별로 보기 좋게 저장한다.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

from models import NonClassSlot, TeacherUnavailable, SimilarSubjectGroup

_HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEAD_FONT = Font(bold=True)


def _write_header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL


def save_settings_xlsx(path, non_class, unavail, similar,
                       year, semester, params, roles=None):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "비수업시간"
    _write_header(ws1, ["학년", "요일", "교시"])
    for s in non_class:
        ws1.append([s.grade, s.day, s.period])
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 8
    ws1.column_dimensions["C"].width = 8

    ws2 = wb.create_sheet("교사불가시간")
    _write_header(ws2, ["교사", "요일", "교시", "학년(0=전체)"])
    for s in unavail:
        ws2.append([s.teacher, s.day, s.period, getattr(s, "grade", 0) or 0])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["D"].width = 12

    ws3 = wb.create_sheet("유사과목그룹")
    _write_header(ws3, ["그룹명", "과목들(쉼표로 구분)"])
    for g in similar:
        ws3.append([g.name, ", ".join(g.subjects)])
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 50

    ws_role = wb.create_sheet("역할")
    _write_header(ws_role, ["교사", "역할"])
    for t, r in (roles or []):
        ws_role.append([t, r])
    ws_role.column_dimensions["A"].width = 14
    ws_role.column_dimensions["B"].width = 12

    ws4 = wb.create_sheet("기본설정")
    _write_header(ws4, ["항목", "값"])
    rows = [
        ("학년도", year),
        ("학기", semester),
        ("교사_최대연속", params.get("max_consecutive", 2)),
        ("하루시수_여유", params.get("daily_n", 1)),
        ("탐색시간_초", params.get("time_limit", 90)),
        ("점심전후_연속방지", 1 if params.get("lunch_split") else 0),
        ("점심_직전교시", params.get("lunch_period", 4)),
    ]
    for r in rows:
        ws4.append(list(r))
    ws4.column_dimensions["A"].width = 20

    wb.save(path)


def _cell(v):
    return v if v is not None else ""


def load_settings_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    non_class, unavail, similar = [], [], []
    settings = {}

    if "비수업시간" in wb.sheetnames:
        for row in wb["비수업시간"].iter_rows(min_row=2, values_only=True):
            if row is None or _cell(row[0]) == "":
                continue
            try:
                non_class.append(NonClassSlot(grade=int(row[0]),
                                              day=str(row[1]).strip(),
                                              period=int(row[2])))
            except (ValueError, TypeError, IndexError):
                continue

    if "교사불가시간" in wb.sheetnames:
        for row in wb["교사불가시간"].iter_rows(min_row=2, values_only=True):
            if row is None or _cell(row[0]) == "":
                continue
            try:
                grade = 0
                if len(row) > 3 and _cell(row[3]) != "":
                    grade = int(row[3])
                unavail.append(TeacherUnavailable(teacher=str(row[0]).strip(),
                                                  day=str(row[1]).strip(),
                                                  period=int(row[2]),
                                                  grade=grade))
            except (ValueError, TypeError, IndexError):
                continue

    if "유사과목그룹" in wb.sheetnames:
        for row in wb["유사과목그룹"].iter_rows(min_row=2, values_only=True):
            if row is None or _cell(row[0]) == "":
                continue
            name = str(row[0]).strip()
            subs = [s.strip() for s in str(_cell(row[1])).split(",") if s.strip()]
            if name and len(subs) >= 2:
                similar.append(SimilarSubjectGroup(name=name, subjects=subs))

    if "기본설정" in wb.sheetnames:
        kv = {}
        for row in wb["기본설정"].iter_rows(min_row=2, values_only=True):
            if row is None or _cell(row[0]) == "":
                continue
            kv[str(row[0]).strip()] = row[1]

        def _int(key, default):
            try:
                return int(kv[key])
            except (KeyError, ValueError, TypeError):
                return default
        settings = {
            "year": _int("학년도", 2026),
            "semester": _int("학기", 1),
            "max_consecutive": _int("교사_최대연속", 2),
            "daily_n": _int("하루시수_여유", 1),
            "time_limit": _int("탐색시간_초", 90),
            "lunch_split": bool(_int("점심전후_연속방지", 0)),
            "lunch_period": _int("점심_직전교시", 4),
        }

    roles = []
    if "역할" in wb.sheetnames:
        for row in wb["역할"].iter_rows(min_row=2, values_only=True):
            if row is None or _cell(row[0]) == "":
                continue
            t = str(row[0]).strip(); r = str(_cell(row[1])).strip()
            if t and r in ("교무부장", "학년부장", "홍보담당"):
                roles.append((t, r))

    return {
        "non_class": non_class,
        "unavail": unavail,
        "similar": similar,
        "settings": settings,
        "roles": roles,
    }
