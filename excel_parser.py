"""학교 시간표 생성기 v3.4.0 - 엑셀 파서"""
from collections import defaultdict
from typing import Optional
import openpyxl
from models import (SchoolData, FixedAssignment, BundleGroup, BundleMember, SpecialRoom)


def _normalize(s) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _is_int_like(s: str) -> bool:
    try:
        int(s)
        return True
    except (ValueError, TypeError):
        return False


def parse_excel(path: str) -> SchoolData:
    """엑셀 파일을 읽어 SchoolData로 변환"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 헤더 분석 (1행: 학년, 2행: 반/특별실)
    grade_header = [_normalize(c.value) for c in ws[1]]
    label_header = [_normalize(c.value) for c in ws[2]]

    # 학년별 컬럼 범위 추출
    grade_cols: dict = {}  # {grade: [(col_idx, label), ...]}
    current_grade: Optional[int] = None
    for i, gh in enumerate(grade_header):
        if "1학년" in gh:
            current_grade = 1
        elif "2학년" in gh:
            current_grade = 2
        elif "3학년" in gh:
            current_grade = 3
        elif "계" in gh:
            current_grade = None
        if current_grade is not None and i < len(label_header):
            lbl = label_header[i]
            if lbl and lbl != "계":
                grade_cols.setdefault(current_grade, []).append((i, lbl))

    # 학급/특별실 분류
    data = SchoolData()
    classes_per_grade: dict = defaultdict(list)
    special_rooms = []
    col_to_classid = {}  # col_idx -> "1-1" 또는 "3-도"

    for grade, cols in grade_cols.items():
        for col_idx, lbl in cols:
            if _is_int_like(lbl):
                cid = f"{grade}-{lbl}"
                classes_per_grade[grade].append(cid)
                col_to_classid[col_idx] = cid
            else:
                cid = f"{grade}-{lbl}"
                special_rooms.append(SpecialRoom(grade=grade, code=lbl))
                col_to_classid[col_idx] = cid

    # 핵심 컬럼 인덱스 찾기 (교사/과목/타임)
    teacher_col = subject_col = time_col = None
    for i, lbl in enumerate(label_header):
        if "교사" in lbl: teacher_col = i
        elif "과목" in lbl: subject_col = i
        elif "타임" in lbl: time_col = i
    if teacher_col is None: teacher_col = 0
    if subject_col is None: subject_col = 1
    if time_col is None: time_col = 2

    # 데이터 행 파싱 (3행부터)
    bundle_raw = defaultdict(list)  # (grade, code) -> [(teacher, subject, class_id, hours)]
    teachers_set = set()
    subjects_set = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        teacher = _normalize(row[teacher_col]) if teacher_col < len(row) else ""
        subject = _normalize(row[subject_col]) if subject_col < len(row) else ""
        time_code = _normalize(row[time_col]) if time_col < len(row) else ""
        if not teacher or not subject:
            continue

        teachers_set.add(teacher)
        subjects_set.add(subject)

        for col_idx, cid in col_to_classid.items():
            if col_idx >= len(row): continue
            v = row[col_idx]
            if v is None or _normalize(v) == "": continue
            try:
                hours = int(v)
            except (ValueError, TypeError):
                continue
            if hours <= 0: continue

            grade = int(cid.split("-")[0])

            if time_code:
                # 묶음수업
                bundle_raw[(grade, time_code)].append((teacher, subject, cid, hours))
            else:
                # 학교지정과목
                data.fixed_assignments.append(
                    FixedAssignment(teacher=teacher, subject=subject, class_id=cid, hours=hours)
                )

    # 묶음 그룹 정리 (v3.3.1 버그 수정 유지: 같은 학급 멤버 hours 합산)
    for (grade, code), rows in bundle_raw.items():
        # 학급별 시수 합산 (분담 처리)
        class_hours_sum = defaultdict(int)
        for teacher, subject, cid, hours in rows:
            class_hours_sum[cid] += hours

        # 묶음 슬롯 수 = 학급별 최댓값
        bg_hours = max(class_hours_sum.values()) if class_hours_sum else 0

        # 멤버는 row 그대로 보존 (h>1인 경우 그대로 hours 유지)
        members = [BundleMember(teacher=t, subject=s, class_id=c, hours=h)
                   for t, s, c, h in rows]

        bg = BundleGroup(grade=grade, code=code, hours=bg_hours, members=members)
        data.bundle_groups.append(bg)

    # 정리
    for g in classes_per_grade:
        classes_per_grade[g].sort(key=lambda x: int(x.split("-")[1]))
    data.classes_per_grade = dict(classes_per_grade)
    data.special_rooms = special_rooms
    data.teachers = sorted(teachers_set)
    data.subjects = sorted(subjects_set)
    return data


def compute_class_total_hours(data: SchoolData) -> dict:
    """학급별 총 시수 계산 (사전진단용)"""
    totals = defaultdict(int)
    for a in data.fixed_assignments:
        totals[a.class_id] += a.hours
    for bg in data.bundle_groups:
        for m in bg.members:
            totals[m.class_id] += m.hours
    return dict(totals)


def compute_teacher_total_hours(data: SchoolData) -> dict:
    """교사별 총 시수 계산"""
    totals = defaultdict(int)
    for a in data.fixed_assignments:
        totals[a.teacher] += a.hours
    for bg in data.bundle_groups:
        for m in bg.members:
            totals[m.teacher] += m.hours
    return dict(totals)
