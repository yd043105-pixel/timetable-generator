"""학교 시간표 생성기 v7.4.0 - 엑셀 출력 (묶음별 색상)"""
import os
from collections import defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["월", "화", "수", "목", "금"]
PERIODS = list(range(1, 8))

# 색상
COLOR_HEADER_MAIN = "1F4E78"   # 진한 청색
COLOR_HEADER_SUB = "BDD7EE"    # 연한 청색
COLOR_NONCLASS = "BFBFBF"      # 회색
COLOR_UNAVAIL = "F4B084"       # 주황
COLOR_SPECIAL = "C6E0B4"       # 초록 (특별실)
COLOR_BUNDLE = "FFF2CC"        # 옅은 노랑 (묶음 기본/예비)

# 묶음별 구분 색상 팔레트 (파스텔 20색)
BUNDLE_PALETTE = [
    "FCE4D6", "FFF2CC", "E2EFDA", "DDEBF7", "FCE4EC", "EDE7F6",
    "FFF9C4", "D7F3E3", "FBE9E7", "E1F5FE", "F3E5F5", "FFF3E0",
    "E8F5E9", "E0F7FA", "FFEBEE", "F1F8E9", "E3F2FD", "FFF8E1",
    "F9FBE7", "EFEBE9",
]


def _bundle_color_map(sol):
    """묶음 코드(bkey)별로 팔레트 색상 1개씩 안정적으로 배정."""
    bkeys = sorted({a[5] for a in sol.assignments if a[5]})
    cmap = {}
    for i, bk in enumerate(bkeys):
        cmap[bk] = BUNDLE_PALETTE[i % len(BUNDLE_PALETTE)]
    return cmap


def _thin_border():
    s = Side(style="thin", color="808080")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_cell(cell, value, *, bold=False, fill=None, align="center", size=10):
    cell.value = value
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.font = Font(name="맑은 고딕", size=size, bold=bold,
                     color="FFFFFF" if fill == COLOR_HEADER_MAIN else "000000")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.border = _thin_border()


def save_excel(sol, data, non_class, unavail, year, semester, params, out_dir="."):
    """결과 시간표를 엑셀로 저장"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 묶음별 색상 맵
    bundle_colors = _bundle_color_map(sol)

    # ── 1. 요약 ─────────────────
    ws = wb.create_sheet("요약")
    _set_cell(ws.cell(row=1, column=1), f"{year}학년도 {semester}학기 시간표 (v7.4.0)",
              bold=True, fill=COLOR_HEADER_MAIN, size=14)
    ws.merge_cells("A1:D1")
    rows = [
        ("생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("최종 페널티", sol.penalty),
        ("배치된 수업 수", len(sol.assignments)),
    ]
    for i, (k, v) in enumerate(rows, 2):
        _set_cell(ws.cell(row=i, column=1), k, bold=True, fill=COLOR_HEADER_SUB)
        _set_cell(ws.cell(row=i, column=2), v)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    # 위반 통계
    r = len(rows) + 3
    _set_cell(ws.cell(row=r, column=1), "위반 통계", bold=True, fill=COLOR_HEADER_SUB)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for k, v in sorted(sol.violations.items()):
        _set_cell(ws.cell(row=r, column=1), k)
        _set_cell(ws.cell(row=r, column=2), v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # 인덱스 구축
    all_classes = sorted(set(a[2] for a in sol.assignments),
                        key=lambda c: (int(c.split("-")[0]), c.split("-")[1]))
    all_teachers = sorted(set(a[1] for a in sol.assignments))

    # 학급 -> {(d,p): (subj, tch, bk)}
    class_table = defaultdict(dict)
    for subj, tch, cid, d, p, bk in sol.assignments:
        class_table[cid][(d, p)] = (subj, tch, bk)

    # 교사 -> {(d,p): (subj, cid, bk)}
    teacher_table = defaultdict(dict)
    for subj, tch, cid, d, p, bk in sol.assignments:
        if (d, p) in teacher_table[tch]:
            # 중복 — 묶음 같이 표시
            prev = teacher_table[tch][(d, p)]
            teacher_table[tch][(d, p)] = (subj, cid + "," + prev[1], bk)
        else:
            teacher_table[tch][(d, p)] = (subj, cid, bk)

    # 비수업
    nonclass_set = set()  # (grade, d, p)
    for s in non_class:
        nonclass_set.add((s.grade, s.day, s.period))
    nc_label = {}
    for s in non_class:
        nc_label[(s.grade, s.day, s.period)] = s.label

    # 불가시간: 교사 시간표의 '불가' 표시는 전체학년(grade=0) 불가일 때만.
    # 학년별 불가는 다른 학년 수업이 그 칸에 올 수 있으므로 '불가'로 막지 않는다.
    unavail_all = defaultdict(set)
    for u in unavail:
        if (getattr(u, "grade", 0) or 0) == 0:
            unavail_all[u.teacher].add((u.day, u.period))

    # 특별실
    special_set = set()
    for sr in data.special_rooms:
        special_set.add(f"{sr.grade}-{sr.code}")

    # ── 2. 학교_전체시간표 ─────────────────
    ws = wb.create_sheet("학교_전체시간표")
    row = 1
    _set_cell(ws.cell(row=row, column=1), f"{year}학년도 {semester}학기 학교 전체 시간표",
              bold=True, fill=COLOR_HEADER_MAIN, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    # 학년별 섹션
    by_grade = defaultdict(list)
    for cid in all_classes:
        by_grade[int(cid.split("-")[0])].append(cid)

    for grade in sorted(by_grade.keys()):
        cls_list = by_grade[grade]
        _set_cell(ws.cell(row=row, column=1), f"{grade}학년 전체 시간표",
                  bold=True, fill=COLOR_HEADER_MAIN, size=12)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=1 + len(cls_list) * 5)
        row += 1
        # 학급 헤더
        _set_cell(ws.cell(row=row, column=1), "교시", bold=True, fill=COLOR_HEADER_SUB)
        col = 2
        for cid in cls_list:
            is_special = cid in special_set
            label = cid + ("★" if is_special else "")
            _set_cell(ws.cell(row=row, column=col), label, bold=True,
                      fill=COLOR_SPECIAL if is_special else COLOR_HEADER_SUB)
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 4)
            col += 5
        row += 1
        # 요일 헤더
        _set_cell(ws.cell(row=row, column=1), "", fill=COLOR_HEADER_SUB)
        col = 2
        for cid in cls_list:
            for d in DAYS:
                _set_cell(ws.cell(row=row, column=col), d, bold=True, fill=COLOR_HEADER_SUB, size=9)
                col += 1
        row += 1
        # 교시별 데이터
        for p in PERIODS:
            _set_cell(ws.cell(row=row, column=1), str(p), bold=True, fill=COLOR_HEADER_SUB)
            col = 2
            for cid in cls_list:
                for d in DAYS:
                    if (grade, d, p) in nonclass_set:
                        _set_cell(ws.cell(row=row, column=col),
                                  nc_label.get((grade, d, p), "비"),
                                  fill=COLOR_NONCLASS, size=9)
                    elif (d, p) in class_table[cid]:
                        subj, tch, bk = class_table[cid][(d, p)]
                        fill = bundle_colors.get(bk, COLOR_BUNDLE) if bk else None
                        _set_cell(ws.cell(row=row, column=col),
                                  f"{subj}\n({tch})", fill=fill, size=9)
                    else:
                        _set_cell(ws.cell(row=row, column=col), "", size=9)
                    col += 1
            row += 1
        row += 1  # 학년 사이 공백

    # 열 너비
    ws.column_dimensions["A"].width = 6
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # ── 3. 교사_전체시간표 ─────────────────
    ws = wb.create_sheet("교사_전체시간표")
    row = 1
    _set_cell(ws.cell(row=row, column=1), f"{year}학년도 {semester}학기 교사 전체 시간표",
              bold=True, fill=COLOR_HEADER_MAIN, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=40)
    row += 2

    # 헤더: 요일/교시 + 시수 + 교사
    _set_cell(ws.cell(row=row, column=1), "", fill=COLOR_HEADER_SUB)
    col = 2
    for d in DAYS:
        _set_cell(ws.cell(row=row, column=col), d, bold=True, fill=COLOR_HEADER_SUB)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 6)
        col += 7
    _set_cell(ws.cell(row=row, column=col), "시수", bold=True, fill=COLOR_HEADER_SUB)
    _set_cell(ws.cell(row=row, column=col + 1), "교사", bold=True, fill=COLOR_HEADER_SUB)
    row += 1
    # 교시 헤더
    _set_cell(ws.cell(row=row, column=1), "교사", bold=True, fill=COLOR_HEADER_SUB)
    col = 2
    for d in DAYS:
        for p in PERIODS:
            _set_cell(ws.cell(row=row, column=col), str(p), bold=True, fill=COLOR_HEADER_SUB, size=9)
            col += 1
    _set_cell(ws.cell(row=row, column=col), "", fill=COLOR_HEADER_SUB)
    _set_cell(ws.cell(row=row, column=col + 1), "", fill=COLOR_HEADER_SUB)
    row += 1

    # 교사별 행
    for tch in all_teachers:
        _set_cell(ws.cell(row=row, column=1), tch, bold=True, size=9)
        total_hours = 0
        col = 2
        for d in DAYS:
            for p in PERIODS:
                if (d, p) in teacher_table[tch]:
                    subj, cids, bk = teacher_table[tch][(d, p)]
                    # 분반 표시: cid 그대로 + 묶음코드 + 학급번호
                    # 예: "3-도" + 묶음 "3_S" → "3s도"
                    parts = []
                    for cid in cids.split(","):
                        g = cid.split("-")[0]
                        cnum = cid.split("-")[1]
                        if bk:
                            bk_code = bk.split("_")[1].lower()
                            parts.append(f"{g}{bk_code}{cnum}")
                        else:
                            parts.append(f"{g}{cnum}")
                    label = "/".join(parts)
                    fill = bundle_colors.get(bk, COLOR_BUNDLE) if bk else None
                    _set_cell(ws.cell(row=row, column=col), label, fill=fill, size=8)
                    total_hours += 1
                elif (d, p) in unavail_all.get(tch, set()):
                    _set_cell(ws.cell(row=row, column=col), "불가",
                              fill=COLOR_UNAVAIL, size=8)
                else:
                    _set_cell(ws.cell(row=row, column=col), "", size=8)
                col += 1
        _set_cell(ws.cell(row=row, column=col), total_hours, bold=True, fill=COLOR_HEADER_SUB)
        _set_cell(ws.cell(row=row, column=col + 1), tch, bold=True, fill=COLOR_HEADER_SUB, size=9)
        row += 1

    # 합계 행 (교시별 운영 학급 수)
    _set_cell(ws.cell(row=row, column=1), "계", bold=True, fill=COLOR_HEADER_SUB)
    col = 2
    grand_total = 0
    for d in DAYS:
        for p in PERIODS:
            cnt = sum(1 for tch in all_teachers if (d, p) in teacher_table[tch])
            _set_cell(ws.cell(row=row, column=col), cnt, bold=True, fill=COLOR_HEADER_SUB, size=9)
            grand_total += cnt
            col += 1
    _set_cell(ws.cell(row=row, column=col), grand_total, bold=True, fill=COLOR_HEADER_SUB)
    _set_cell(ws.cell(row=row, column=col + 1), "", fill=COLOR_HEADER_SUB)

    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6

    # ── 4. 학급별 개별 시트 ─────────────────
    for cid in all_classes:
        sheet_name = cid.replace("/", "_")[:31]
        ws = wb.create_sheet(sheet_name)
        is_special = cid in special_set
        title = f"{cid} 시간표" + (" (특별실)" if is_special else "")
        _set_cell(ws.cell(row=1, column=1), title,
                  bold=True, fill=COLOR_HEADER_MAIN, size=14)
        ws.merge_cells("A1:F1")
        # 헤더
        _set_cell(ws.cell(row=2, column=1), "교시", bold=True, fill=COLOR_HEADER_SUB)
        for i, d in enumerate(DAYS, 2):
            _set_cell(ws.cell(row=2, column=i), d, bold=True, fill=COLOR_HEADER_SUB)
        grade = int(cid.split("-")[0])
        for p in PERIODS:
            _set_cell(ws.cell(row=p + 2, column=1), str(p), bold=True, fill=COLOR_HEADER_SUB)
            for i, d in enumerate(DAYS, 2):
                if (grade, d, p) in nonclass_set:
                    _set_cell(ws.cell(row=p + 2, column=i),
                              nc_label.get((grade, d, p), "비"),
                              fill=COLOR_NONCLASS)
                elif (d, p) in class_table[cid]:
                    subj, tch, bk = class_table[cid][(d, p)]
                    fill = bundle_colors.get(bk, COLOR_BUNDLE) if bk else None
                    _set_cell(ws.cell(row=p + 2, column=i), f"{subj}\n({tch})", fill=fill)
                else:
                    _set_cell(ws.cell(row=p + 2, column=i), "")
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # ── 5. 교사별 개별 시트 ─────────────────
    for tch in all_teachers:
        sheet_name = tch.replace("/", "_")[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name + "_T"
        ws = wb.create_sheet(sheet_name)
        _set_cell(ws.cell(row=1, column=1), f"{tch} 시간표",
                  bold=True, fill=COLOR_HEADER_MAIN, size=14)
        ws.merge_cells("A1:F1")
        _set_cell(ws.cell(row=2, column=1), "교시", bold=True, fill=COLOR_HEADER_SUB)
        for i, d in enumerate(DAYS, 2):
            _set_cell(ws.cell(row=2, column=i), d, bold=True, fill=COLOR_HEADER_SUB)
        for p in PERIODS:
            _set_cell(ws.cell(row=p + 2, column=1), str(p), bold=True, fill=COLOR_HEADER_SUB)
            for i, d in enumerate(DAYS, 2):
                if (d, p) in teacher_table[tch]:
                    subj, cids, bk = teacher_table[tch][(d, p)]
                    fill = bundle_colors.get(bk, COLOR_BUNDLE) if bk else None
                    _set_cell(ws.cell(row=p + 2, column=i), f"{subj}\n({cids})", fill=fill)
                elif (d, p) in unavail_all.get(tch, set()):
                    _set_cell(ws.cell(row=p + 2, column=i), "불가", fill=COLOR_UNAVAIL)
                else:
                    _set_cell(ws.cell(row=p + 2, column=i), "")
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # 저장
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"시간표_{year}_{semester}_{ts}.xlsx"
    fpath = os.path.join(out_dir, fname)
    wb.save(fpath)
    return fpath
